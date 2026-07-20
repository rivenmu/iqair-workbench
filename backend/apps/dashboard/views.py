from datetime import datetime

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.audit.mixins import AuditLogMixin
from apps.accounts.permissions import IsAdminOrReadOnly

from .models import Brand, FilterRevenue, PlatformSalesData, Platform, PeriodType
from .serializers import (
    BrandSerializer, BrandListSerializer,
    PlatformSalesDataSerializer,
)
from services.dashboard_service import DashboardService
from services.platform_data_service import PlatformDataService


class BrandViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """品牌 ViewSet"""
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        return Brand.objects.filter(project_id=self.kwargs['project_pk']).order_by('sort_order')

    def get_serializer_class(self):
        if self.action == 'list':
            return BrandListSerializer
        return BrandSerializer

    def perform_create(self, serializer):
        serializer.save(project_id=self.kwargs['project_pk'])


class DashboardViewSet(viewsets.ViewSet):
    """看板数据聚合接口"""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='data')
    def get_data(self, request, project_pk=None):
        """获取完整看板数据"""
        data = DashboardService.get_dashboard_data(project_pk)
        return Response(data)

    @action(detail=False, methods=['post'], url_path='data')
    def save_data(self, request, project_pk=None):
        """保存完整看板数据（批量更新）"""
        try:
            DashboardService.save_dashboard_data(project_pk, request.data, request.user)
            return Response({'detail': '保存成功'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='ui-texts')
    def save_ui_texts(self, request, project_pk=None):
        """
        轻量保存 UI 文本（标题/表头等可编辑文案）
        用于标题失焦等场景，避免触发全量品牌数据写入
        请求体：{ "uiTexts": { "mainTitle": "...", "thSales": "..." } }
        """
        try:
            ui_texts = request.data.get('uiTexts', {}) or {}
            if not isinstance(ui_texts, dict):
                return Response(
                    {'detail': 'uiTexts 必须是对象'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            DashboardService.save_ui_texts(project_pk, ui_texts)
            return Response({'detail': 'UI 文本已保存', 'uiTexts': ui_texts}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class PlatformDataViewSet(viewsets.ViewSet):
    """电商平台数据接口（上传 + 查询）"""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @action(detail=False, methods=['post'], url_path='upload')
    def upload_excel(self, request):
        """上传 Excel 文件并解析入库"""
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请选择文件'}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': '仅支持 .xlsx / .xls 格式'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            file_bytes = file.read()
            parsed = PlatformDataService.parse_excel(file_bytes)
            if not parsed['platforms']:
                return Response(
                    {'detail': f'未解析到有效数据。{parsed["summary"]}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            saved = PlatformDataService.save_parsed_data(parsed, request.user)
            return Response({
                'detail': f'上传成功，共保存 {saved} 条记录',
                'summary': parsed['summary'],
                'saved_count': saved,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': f'解析失败: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='query')
    def query_data(self, request):
        """查询平台数据（含同比）"""
        platform = request.query_params.get('platform', Platform.TMALL)
        period_type = request.query_params.get('period_type', PeriodType.DAILY)
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if platform not in dict(Platform.choices):
            return Response({'detail': '无效的平台参数'}, status=status.HTTP_400_BAD_REQUEST)
        if period_type not in dict(PeriodType.choices):
            return Response({'detail': '无效的周期类型'}, status=status.HTTP_400_BAD_REQUEST)

        sd = None
        ed = None
        if start_date:
            try:
                sd = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'start_date 格式应为 YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
        if end_date:
            try:
                ed = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'end_date 格式应为 YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

        data = PlatformDataService.query_data(platform, period_type, sd, ed)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='range')
    def date_range(self, request):
        """获取可用日期范围"""
        platform = request.query_params.get('platform', Platform.TMALL)
        period_type = request.query_params.get('period_type', PeriodType.DAILY)
        data = PlatformDataService.get_available_date_range(platform, period_type)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        """下载 Excel 导入模板"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = Workbook()
        # 天猫 sheet
        ws_tm = wb.active
        ws_tm.title = '天猫'
        ws_tm.append(['日期', '销售额', '订单数', '访客数', '支付买家数', '支付转化率', '客单价', '加购人数', '收藏人数'])
        ws_tm.append(['2026-07-01', 158000, 320, 8500, 280, 3.29, 493.75, 1200, 600])
        ws_tm.append(['2026-07-02', 165000, 340, 8800, 295, 3.35, 485.29, 1280, 640])
        # 京东 sheet
        ws_jd = wb.create_sheet('京东')
        ws_jd.append(['日期', '销售额', '订单数', '访客数', '支付买家数', '支付转化率', '客单价', '加购人数', '收藏人数'])
        ws_jd.append(['2026-07-01', 132000, 260, 7200, 220, 3.06, 507.69, 980, 480])
        ws_jd.append(['2026-07-02', 138000, 275, 7500, 232, 3.09, 501.82, 1020, 510])

        for ws in [ws_tm, ws_jd]:
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 14

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="platform_data_template.xlsx"'
        wb.save(response)
        return response

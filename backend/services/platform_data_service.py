"""电商平台数据服务层 — Excel 解析 + 查询（含同比）"""
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db.models import Min, Max, Count

from openpyxl import load_workbook

from apps.dashboard.models import PlatformSalesData, Platform, PeriodType


# Excel 表头中文名 → 模型字段的映射
COLUMN_MAP = {
    '日期': 'date',
    '销售额': 'sales_amount',
    '销售金额': 'sales_amount',
    '支付金额': 'sales_amount',
    '成交金额': 'sales_amount',
    '订单数': 'order_count',
    '订单量': 'order_count',
    '访客数': 'visitor_count',
    '支付买家数': 'paying_buyer_count',
    '支付转化率': 'conversion_rate',
    '转化率': 'conversion_rate',
    '客单价': 'unit_price',
    '加购人数': 'cart_count',
    '收藏人数': 'favorite_count',
}

# sheet 名 → platform 的映射
SHEET_MAP = {
    '天猫': Platform.TMALL,
    'tmall': Platform.TMALL,
    '生意参谋': Platform.TMALL,
    '京东': Platform.JD,
    'jd': Platform.JD,
    '京东商智': Platform.JD,
}


def _parse_decimal(val) -> Decimal:
    if val is None:
        return Decimal('0')
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '').replace('%', '').replace('¥', '')
    if not s or s == '-':
        return Decimal('0')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _parse_int(val) -> int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(',', '')
    if not s or s == '-':
        return 0
    try:
        return int(float(s))
    except (InvalidOperation, ValueError):
        return 0


def _parse_date(val):
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class PlatformDataService:
    """电商平台数据服务"""

    @staticmethod
    def parse_excel(file_bytes: bytes) -> dict:
        """
        解析上传的 Excel 文件。

        Excel 格式约定：
        - 每个 sheet 对应一个平台（sheet 名含"天猫"/"京东"或英文 tmall/jd）
        - 每个数据 sheet 的第一行周期类型标记（可选）：日/周/月，否则默认 daily
          实际数据从含"日期"表头的那一行开始
        - 必须包含"日期"列，其他指标列可选

        返回: {'platforms': {platform: {period_type: [row_dict, ...]}}, 'summary': str}
        """
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        result = {}
        summary_parts = []

        for sheet_name in wb.sheetnames:
            platform = None
            for key, plat in SHEET_MAP.items():
                if key.lower() in sheet_name.lower():
                    platform = plat
                    break
            if not platform:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 查找表头行（包含"日期"的行）
            header_row_idx = None
            for idx, row in enumerate(rows[:5]):
                if row and any(cell and '日期' in str(cell) for cell in row):
                    header_row_idx = idx
                    break

            if header_row_idx is None:
                summary_parts.append(f'[{sheet_name}] 未找到"日期"表头，跳过')
                continue

            header_row = rows[header_row_idx]
            col_mapping = {}
            for col_idx, cell in enumerate(header_row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                if cell_str in COLUMN_MAP:
                    col_mapping[col_idx] = COLUMN_MAP[cell_str]

            if 'date' not in col_mapping.values():
                summary_parts.append(f'[{sheet_name}] 表头缺少"日期"列，跳过')
                continue

            # 检查表头上方是否有周期类型标记
            period_type = PeriodType.DAILY
            if header_row_idx > 0:
                for cell in rows[header_row_idx - 1]:
                    if cell:
                        cell_str = str(cell).strip()
                        if '周' == cell_str or 'weekly' in cell_str.lower():
                            period_type = PeriodType.WEEKLY
                            break
                        elif '月' == cell_str or 'monthly' in cell_str.lower():
                            period_type = PeriodType.MONTHLY
                            break
                        elif '日' == cell_str or 'daily' in cell_str.lower():
                            period_type = PeriodType.DAILY
                            break

            data_rows = []
            for row in rows[header_row_idx + 1:]:
                if not row or all(c is None for c in row):
                    continue
                row_data = {}
                for col_idx, field in col_mapping.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    if field == 'date':
                        parsed = _parse_date(val)
                        if parsed:
                            row_data['date'] = parsed
                    elif field in ('sales_amount', 'conversion_rate', 'unit_price'):
                        row_data[field] = _parse_decimal(val)
                    else:
                        row_data[field] = _parse_int(val)

                if 'date' in row_data:
                    data_rows.append(row_data)

            if platform not in result:
                result[platform] = {}
            result[platform][period_type] = data_rows
            summary_parts.append(f'[{sheet_name}] 解析 {len(data_rows)} 条记录 ({period_type})')

        return {
            'platforms': result,
            'summary': '; '.join(summary_parts) if summary_parts else '未解析到有效数据',
        }

    @staticmethod
    def save_parsed_data(parsed: dict, user=None) -> int:
        """保存解析后的数据到数据库（upsert）"""
        saved_count = 0
        for platform, period_data in parsed['platforms'].items():
            for period_type, rows in period_data.items():
                for row in rows:
                    obj, created = PlatformSalesData.objects.update_or_create(
                        platform=platform,
                        period_type=period_type,
                        date=row['date'],
                        defaults={
                            'period_label': PlatformDataService._build_period_label(row['date'], period_type),
                            'sales_amount': row.get('sales_amount', Decimal('0')),
                            'order_count': row.get('order_count', 0),
                            'visitor_count': row.get('visitor_count', 0),
                            'paying_buyer_count': row.get('paying_buyer_count', 0),
                            'conversion_rate': row.get('conversion_rate', Decimal('0')),
                            'unit_price': row.get('unit_price', Decimal('0')),
                            'cart_count': row.get('cart_count', 0),
                            'favorite_count': row.get('favorite_count', 0),
                            'uploaded_by': user,
                        }
                    )
                    saved_count += 1
        return saved_count

    @staticmethod
    def _build_period_label(d: date, period_type: str) -> str:
        if period_type == PeriodType.DAILY:
            return d.strftime('%Y-%m-%d')
        elif period_type == PeriodType.WEEKLY:
            iso = d.isocalendar()
            return f'{iso.year}-W{iso.week:02d}'
        else:
            return d.strftime('%Y-%m')

    @staticmethod
    def query_data(
        platform: str,
        period_type: str,
        start_date=None,
        end_date=None,
    ) -> dict:
        """
        查询平台数据，自动附带去年同期同比。
        """
        qs = PlatformSalesData.objects.filter(
            platform=platform, period_type=period_type
        )
        if start_date:
            qs = qs.filter(date__gte=start_date)
        if end_date:
            qs = qs.filter(date__lte=end_date)

        records = list(qs.order_by('date'))

        platform_display = dict(Platform.choices).get(platform, platform)

        if not records:
            return {
                'platform': platform,
                'platform_display': platform_display,
                'period_type': period_type,
                'records': [],
                'summary': {},
            }

        # 构建去年同期数据映射（批量查询减少 N+1）
        yoy_dates = []
        for r in records:
            try:
                yoy_dates.append(r.date.replace(year=r.date.year - 1))
            except ValueError:
                pass  # 2/29 等特殊日期

        yoy_qs = PlatformSalesData.objects.filter(
            platform=platform, period_type=period_type, date__in=yoy_dates
        )
        yoy_map = {r.date: r for r in yoy_qs}

        result_records = []
        for r in records:
            try:
                yoy_date = r.date.replace(year=r.date.year - 1)
            except ValueError:
                yoy_date = None
            yoy = yoy_map.get(yoy_date) if yoy_date else None

            rec = {
                'date': r.date.isoformat(),
                'label': r.period_label or r.date.isoformat(),
                'sales_amount': float(r.sales_amount),
                'order_count': r.order_count,
                'visitor_count': r.visitor_count,
                'paying_buyer_count': r.paying_buyer_count,
                'conversion_rate': float(r.conversion_rate),
                'unit_price': float(r.unit_price),
                'cart_count': r.cart_count,
                'favorite_count': r.favorite_count,
            }
            if yoy:
                rec['yoy_sales_amount'] = float(yoy.sales_amount)
                rec['yoy_order_count'] = yoy.order_count
                rec['yoy_visitor_count'] = yoy.visitor_count
                rec['yoy_conversion_rate'] = float(yoy.conversion_rate)
                rec['yoy_growth'] = round(
                    float((r.sales_amount - yoy.sales_amount) / yoy.sales_amount * 100), 2
                ) if yoy.sales_amount and yoy.sales_amount > 0 else None
            else:
                rec['yoy_sales_amount'] = None
                rec['yoy_order_count'] = None
                rec['yoy_visitor_count'] = None
                rec['yoy_conversion_rate'] = None
                rec['yoy_growth'] = None

            result_records.append(rec)

        # 汇总
        total_sales = sum(r['sales_amount'] for r in result_records)
        total_orders = sum(r['order_count'] for r in result_records)
        total_visitors = sum(r['visitor_count'] for r in result_records)
        avg_conversion = (
            sum(r['conversion_rate'] for r in result_records) / len(result_records)
            if result_records else 0
        )
        yoy_total_sales = sum(
            r['yoy_sales_amount'] for r in result_records if r['yoy_sales_amount'] is not None
        )
        yoy_growth = None
        if yoy_total_sales and yoy_total_sales > 0:
            yoy_growth = round((total_sales - yoy_total_sales) / yoy_total_sales * 100, 2)

        return {
            'platform': platform,
            'platform_display': platform_display,
            'period_type': period_type,
            'records': result_records,
            'summary': {
                'total_sales': round(total_sales, 2),
                'total_orders': total_orders,
                'total_visitors': total_visitors,
                'avg_conversion': round(avg_conversion, 2),
                'yoy_total_sales': round(yoy_total_sales, 2) if yoy_total_sales else None,
                'yoy_growth': yoy_growth,
                'record_count': len(result_records),
            },
        }

    @staticmethod
    def get_available_date_range(platform: str, period_type: str) -> dict:
        """获取指定平台+周期的可用日期范围"""
        qs = PlatformSalesData.objects.filter(platform=platform, period_type=period_type)
        if not qs.exists():
            return {'min_date': None, 'max_date': None, 'count': 0}
        agg = qs.aggregate(
            min_d=Min('date'),
            max_d=Max('date'),
            cnt=Count('id'),
        )
        return {
            'min_date': agg['min_d'].isoformat() if agg['min_d'] else None,
            'max_date': agg['max_d'].isoformat() if agg['max_d'] else None,
            'count': agg['cnt'],
        }
